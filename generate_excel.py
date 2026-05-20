#!/usr/bin/env python3
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Leer el README
readme_path = Path(__file__).parent / "README.md"
with open(readme_path, "r", encoding="utf-8") as f:
    content = f.read()

# Crear workbook
wb = Workbook()
wb.remove(wb.active)  # Eliminar hoja por defecto

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
subheader_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def add_header_row(ws, headers):
    """Agrega fila de encabezado con formato"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

def categorize_requirement(title, req_id):
    """Categoriza requerimiento basado en título e ID"""
    title_lower = title.lower()
    
    if "health" in title_lower or "salud" in title_lower:
        return "Health Check"
    elif "registro" in title_lower or "login" in title_lower or "refresh" in title_lower or "logout" in title_lower or "autenticac" in title_lower or "correo" in title_lower or "otp" in title_lower:
        return "Autenticación"
    elif "perfil" in title_lower or "usuario" in title_lower and "public" not in title_lower:
        return "Perfil de Usuario"
    elif "vehiculo" in title_lower or "vehículo" in title_lower:
        return "Vehículos"
    elif "viaje" in title_lower or "trip" in title_lower:
        return "Viajes"
    elif "solicitud" in title_lower or "request" in title_lower or "unirse" in title_lower:
        return "Solicitudes"
    elif "pago" in title_lower or "payment" in title_lower or "stripe" in title_lower or "intento" in title_lower:
        return "Pagos"
    elif "calificaci" in title_lower or "rating" in title_lower:
        return "Calificaciones"
    elif "reporte" in title_lower or "report" in title_lower:
        return "Reportes"
    elif "admin" in title_lower:
        return "Administración"
    elif "gps" in title_lower or "tiempo real" in title_lower or "socket" in title_lower or "ubicaci" in title_lower:
        return "GPS/Tiempo Real"
    elif "seguridad" in title_lower or "autenticaci" in title_lower or "jwt" in title_lower or "rol" in title_lower:
        return "Seguridad"
    elif "validaci" in title_lower:
        return "Validación"
    elif "privacidad" in title_lower or "datos sensibles" in title_lower:
        return "Privacidad"
    elif "idempotet" in title_lower:
        return "Confiabilidad"
    elif "rendimiento" in title_lower or "performance" in title_lower:
        return "Rendimiento"
    elif "disponibilidad" in title_lower or "availability" in title_lower:
        return "Disponibilidad"
    elif "mantenibilidad" in title_lower or "testabilidad" in title_lower:
        return "Mantenibilidad"
    elif "usabilidad" in title_lower:
        return "Usabilidad"
    elif "escalabilidad" in title_lower:
        return "Escalabilidad"
    else:
        return "Otros"

def parse_requirements(content, section_type):
    """Parsea requerimientos del README"""
    pattern = rf"### {section_type}-(\d+)\s*-\s*(.+?)\n(.+?)(?=### {section_type}-|\Z)"
    matches = re.findall(pattern, content, re.DOTALL | re.MULTILINE)
    
    results = []
    for match in matches:
        req_id = f"{section_type}-{match[0]}"
        title = match[1].strip()
        description = match[2].strip()
        
        # Extraer casos de prueba
        test_pattern = r"Casos de prueba:\n(.*?)(?=### |$)"
        test_match = re.search(test_pattern, description, re.DOTALL)
        
        test_cases = []
        if test_match:
            test_text = test_match.group(1)
            # Buscar líneas que comienzan con número.
            test_lines = re.findall(r"^\d+\.\s+(.+?)(?=\n\d+\.|$)", test_text, re.MULTILINE | re.DOTALL)
            for test in test_lines:
                test_cases.append(test.strip())
        
        # Extraer descripción sin casos de prueba
        desc_clean = re.sub(r"Casos de prueba:.*", "", description, flags=re.DOTALL).strip()
        
        # Categorizar
        category = categorize_requirement(title, req_id)
        
        results.append({
            'id': req_id,
            'title': title,
            'description': desc_clean,
            'test_cases': test_cases,
            'category': category
        })
    
    return results

# Parsear requerimientos funcionales
fr_list = parse_requirements(content, "FR")
# Parsear requerimientos no funcionales
nfr_list = parse_requirements(content, "NFR")

# ==================== HOJA 1: Resumen de todos los Requerimientos ====================
ws_summary = wb.create_sheet("Requerimientos")
add_header_row(ws_summary, ["ID", "Tipo", "Categoría", "Título", "Descripción", "Casos de Prueba", "Estado"])
ws_summary.column_dimensions["A"].width = 12
ws_summary.column_dimensions["B"].width = 12
ws_summary.column_dimensions["C"].width = 18
ws_summary.column_dimensions["D"].width = 35
ws_summary.column_dimensions["E"].width = 40
ws_summary.column_dimensions["F"].width = 15
ws_summary.column_dimensions["G"].width = 12

row = 2
for req in fr_list:
    ws_summary.cell(row=row, column=1).value = req['id']
    ws_summary.cell(row=row, column=2).value = "Funcional"
    ws_summary.cell(row=row, column=3).value = req['category']
    ws_summary.cell(row=row, column=4).value = req['title']
    ws_summary.cell(row=row, column=5).value = req['description']
    ws_summary.cell(row=row, column=6).value = len(req['test_cases'])
    ws_summary.cell(row=row, column=7).value = "Pendiente"
    
    for col in range(1, 8):
        ws_summary.cell(row=row, column=col).border = border
        ws_summary.cell(row=row, column=col).alignment = left_alignment
    
    row += 1

for req in nfr_list:
    ws_summary.cell(row=row, column=1).value = req['id']
    ws_summary.cell(row=row, column=2).value = "No Funcional"
    ws_summary.cell(row=row, column=3).value = req['category']
    ws_summary.cell(row=row, column=4).value = req['title']
    ws_summary.cell(row=row, column=5).value = req['description']
    ws_summary.cell(row=row, column=6).value = len(req['test_cases'])
    ws_summary.cell(row=row, column=7).value = "Pendiente"
    
    # Colorear filas NFR diferente
    for col in range(1, 8):
        ws_summary.cell(row=row, column=col).border = border
        ws_summary.cell(row=row, column=col).alignment = left_alignment
        ws_summary.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    row += 1

# ==================== HOJA 2: Casos de Prueba Funcionales ====================
ws_fr_tests = wb.create_sheet("FR Test Cases")
add_header_row(ws_fr_tests, ["ID Requerimiento", "Categoría", "Título Requerimiento", "# Caso", "Descripción del Caso", "Estado", "Ejecutado"])
ws_fr_tests.column_dimensions["A"].width = 12
ws_fr_tests.column_dimensions["B"].width = 18
ws_fr_tests.column_dimensions["C"].width = 35
ws_fr_tests.column_dimensions["D"].width = 8
ws_fr_tests.column_dimensions["E"].width = 55
ws_fr_tests.column_dimensions["F"].width = 12
ws_fr_tests.column_dimensions["G"].width = 12

row = 2
for req in fr_list:
    for idx, test in enumerate(req['test_cases'], 1):
        ws_fr_tests.cell(row=row, column=1).value = req['id']
        ws_fr_tests.cell(row=row, column=2).value = req['category']
        ws_fr_tests.cell(row=row, column=3).value = req['title']
        ws_fr_tests.cell(row=row, column=4).value = idx
        ws_fr_tests.cell(row=row, column=5).value = test
        ws_fr_tests.cell(row=row, column=6).value = "Pendiente"
        ws_fr_tests.cell(row=row, column=7).value = ""
        
        for col in range(1, 8):
            ws_fr_tests.cell(row=row, column=col).border = border
            if col == 5:
                ws_fr_tests.cell(row=row, column=col).alignment = left_alignment
            else:
                ws_fr_tests.cell(row=row, column=col).alignment = center_alignment
        
        row += 1

# ==================== HOJA 3: Casos de Prueba No Funcionales ====================
ws_nfr_tests = wb.create_sheet("NFR Test Cases")
add_header_row(ws_nfr_tests, ["ID Requerimiento", "Categoría", "Título Requerimiento", "# Caso", "Descripción del Caso", "Estado", "Ejecutado"])
ws_nfr_tests.column_dimensions["A"].width = 12
ws_nfr_tests.column_dimensions["B"].width = 18
ws_nfr_tests.column_dimensions["C"].width = 35
ws_nfr_tests.column_dimensions["D"].width = 8
ws_nfr_tests.column_dimensions["E"].width = 55
ws_nfr_tests.column_dimensions["F"].width = 12
ws_nfr_tests.column_dimensions["G"].width = 12

row = 2
for req in nfr_list:
    for idx, test in enumerate(req['test_cases'], 1):
        ws_nfr_tests.cell(row=row, column=1).value = req['id']
        ws_nfr_tests.cell(row=row, column=2).value = req['category']
        ws_nfr_tests.cell(row=row, column=3).value = req['title']
        ws_nfr_tests.cell(row=row, column=4).value = idx
        ws_nfr_tests.cell(row=row, column=5).value = test
        ws_nfr_tests.cell(row=row, column=6).value = "Pendiente"
        ws_nfr_tests.cell(row=row, column=7).value = ""
        
        for col in range(1, 8):
            ws_nfr_tests.cell(row=row, column=col).border = border
            if col == 5:
                ws_nfr_tests.cell(row=row, column=col).alignment = left_alignment
            else:
                ws_nfr_tests.cell(row=row, column=col).alignment = center_alignment
        
        row += 1

# ==================== HOJA 4: Matriz de Trazabilidad ====================
ws_matrix = wb.create_sheet("Trazabilidad")
all_reqs = fr_list + nfr_list
total_tests = sum(len(r['test_cases']) for r in all_reqs)

add_header_row(ws_matrix, ["ID", "Tipo", "Categoría", "Título", "Total Casos", "Casos Ejecutados", "% Cobertura", "Estatus"])
ws_matrix.column_dimensions["A"].width = 12
ws_matrix.column_dimensions["B"].width = 12
ws_matrix.column_dimensions["C"].width = 18
ws_matrix.column_dimensions["D"].width = 35
ws_matrix.column_dimensions["E"].width = 15
ws_matrix.column_dimensions["F"].width = 18
ws_matrix.column_dimensions["G"].width = 12
ws_matrix.column_dimensions["H"].width = 12

row = 2
for req in all_reqs:
    req_type = "FR" if req['id'].startswith("FR") else "NFR"
    ws_matrix.cell(row=row, column=1).value = req['id']
    ws_matrix.cell(row=row, column=2).value = req_type
    ws_matrix.cell(row=row, column=3).value = req['category']
    ws_matrix.cell(row=row, column=4).value = req['title']
    ws_matrix.cell(row=row, column=5).value = len(req['test_cases'])
    ws_matrix.cell(row=row, column=6).value = 0
    ws_matrix.cell(row=row, column=7).value = "=F{}/E{}".format(row, row)
    ws_matrix.cell(row=row, column=8).value = "No Iniciado"
    
    for col in range(1, 9):
        ws_matrix.cell(row=row, column=col).border = border
        ws_matrix.cell(row=row, column=col).alignment = center_alignment
    
    # Formatear columna de porcentaje como porcentaje
    ws_matrix.cell(row=row, column=7).number_format = '0%'
    
    row += 1

# Agregar fila de totales
ws_matrix.cell(row=row, column=1).value = "TOTAL"
ws_matrix.cell(row=row, column=1).font = Font(bold=True)
ws_matrix.cell(row=row, column=5).value = f"=SUM(E2:E{row-1})"
ws_matrix.cell(row=row, column=5).font = Font(bold=True)
ws_matrix.cell(row=row, column=6).value = f"=SUM(F2:F{row-1})"
ws_matrix.cell(row=row, column=6).font = Font(bold=True)
ws_matrix.cell(row=row, column=7).value = f"=F{row}/E{row}"
ws_matrix.cell(row=row, column=7).font = Font(bold=True)
ws_matrix.cell(row=row, column=7).number_format = '0%'

for col in range(1, 9):
    ws_matrix.cell(row=row, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws_matrix.cell(row=row, column=col).border = border

# Guardar workbook
import os
output_filename = "REQUERIMIENTOS_PRUEBAS.xlsx"
output_path = Path(__file__).parent / output_filename
# Si el archivo existe, eliminarlo
if output_path.exists():
    try:
        os.remove(output_path)
    except PermissionError:
        # Si está bloqueado, usar un nombre temporal
        output_filename = "REQUERIMIENTOS_PRUEBAS_temp.xlsx"
        output_path = Path(__file__).parent / output_filename

wb.save(output_path)
print(f"✓ Excel generado: {output_path}")
print(f"  - {len(fr_list)} Requerimientos Funcionales")
print(f"  - {len(nfr_list)} Requerimientos No Funcionales")
print(f"  - {total_tests} Casos de prueba totales")
